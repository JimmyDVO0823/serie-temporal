import os
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
from statsmodels.tsa.stattools import acf

AZUL = "#1A3A6B"
ROJO = "#C0392B"
VERDE = "#27AE60"
NARANJ = "#E67E22"
GRIS = "#7F8C8D"

MAPPING_ECONO = {
    "intercept": "Intercepto / Drift (c)",
    "const": "Intercepto / Drift (c)",
    "ar.L1": "phi1  (AR regular lag 1 - Y_t-1)",
    "ar.L2": "phi2  (AR regular lag 2)",
    "ma.L1": "theta1 (MA regular lag 1)",
    "ar.S.L4": "Phi1 (SAR estacional lag 4 - Y_t-4)",
    "ma.S.L4": "Theta1 (SMA estacional lag 4)",
    "sigma2": "sigma2 (Varianza del error)",
}

def generate_text_report(out_dir, variable_name, file_path, adf, acf_pacf, results, validos, mejor, future_labels, p_mu, p_ci):
    report_path = os.path.join(out_dir, "reporte_modelos.txt")
    LINEA = "=" * 60
    lim95 = acf_pacf["lim95"]
    acf_vals = acf_pacf["acf_vals"]
    pacf_vals = acf_pacf["pacf_vals"]

    with open(report_path, "w", encoding="utf-8") as f:
        f.write(LINEA + "\n")
        f.write(f"  REPORTE SARIMA ESTACIONAL — {variable_name.upper()}\n")
        f.write(f"  Archivo: {file_path}\n")
        f.write("  Estructura Completa: [Y_t, Y_t-1, d=1, Y_t-4, D=1]\n")
        f.write(LINEA + "\n\n")

        f.write("── PRUEBA DE ESTACIONARIEDAD (ADF) ──────────────────────\n")
        f.write(f"Serie Yt (Original): estadístico ADF = {adf['y'][0]:.4f} | p-valor = {adf['y'][1]:.4f}\n")
        f.write(f"  → {'Estacionaria' if adf['estac_y'] else 'NO estacionaria → Requiere diferencia regular (d=1)'}\n\n")
        f.write(f"Serie Wt (Diff Regular): estadístico ADF = {adf['wt'][0]:.4f} | p-valor = {adf['wt'][1]:.4f}\n")
        f.write(f"  → {'Estacionaria ✓ (d=1 confirmado)' if adf['estac_wt'] else 'Sugerencia: Revisar combinación con D=1'}\n\n")

        f.write("── ACF / PACF DE Wt (límite ±{:.3f}) ──────────────────\n".format(lim95))
        f.write(f"  {'Lag':<5} {'ACF':>8} {'PACF':>8}  Significativo\n")
        for i in range(len(acf_vals)):
            sig = "SIG" if abs(acf_vals[i]) > lim95 or abs(pacf_vals[i]) > lim95 else "ns"
            f.write(f"  {i+1:<5} {acf_vals[i]:>+8.4f} {pacf_vals[i]:>+8.4f}  {sig}\n")
        f.write("\n")

        f.write(LINEA + "\n")
        f.write("  ESTIMACIÓN DE MODELOS CANDIDATOS\n")
        f.write(LINEA + "\n\n")

        for r in results:
            if r.get("res") is None:
                f.write(f">> {r['nombre']}  —  ERROR: {r.get('error','?')}\n\n")
                continue
            res = r["res"]
            alerta = "  ⚠ Unidades cerca del límite (descartado por invertibilidad)" if not r["ma_ok"] else ""
            f.write(f">> {r['nombre']}{alerta}\n")
            f.write("-" * 50 + "\n")
            f.write(f"  AIC  = {r['aic']:.4f}\n")
            f.write(f"  AICc = {r['aicc']:.4f}\n")
            f.write(f"  BIC  = {r['bic']:.4f}\n")
            f.write(f"  RMSE = {r['rmse']:.4f}\n")
            f.write(f"  MAE  = {r['mae']:.4f}\n")
            f.write(f"  Ljung-Box p-valor = {r['p_lb']:.4f}\n")
            f.write("\n  Parámetros Estimados:\n")
            f.write(f"  {'Parámetro':<22} {'Coeficiente':>12} {'p-valor':>10}  Signif.\n")
            f.write("  " + "-" * 55 + "\n")
            for pname in res.param_names:
                if pname == "sigma2":
                    continue
                coef = res.params[res.param_names.index(pname)]
                pval = res.pvalues[res.param_names.index(pname)]
                disp = MAPPING_ECONO.get(pname, pname)
                sig = "✓ Signif." if pval < 0.05 else "✗ No signif."
                f.write(f"  {disp:<22} {coef:>12.6f} {pval:>10.4f}  {sig}\n")
            sigma2 = res.params[res.param_names.index("sigma2")]
            f.write(f"  {'sigma2 (varianza error)':<22} {sigma2:>12.6f}\n\n")

        f.write(LINEA + "\n")
        f.write("  SELECCIÓN DEL MEJOR MODELO ECONÓMETRICO\n")
        f.write(LINEA + "\n\n")
        f.write("  Criterio de decisión: Mínimo AICc (corregido para muestras pequeñas).\n\n")
        f.write("  Ranking General:\n")
        for i, r in enumerate(sorted(validos, key=lambda x: x["aicc"]), 1):
            marca = " ← SELECCIONADO" if r["nombre"] == mejor["nombre"] else ""
            f.write(f"    {i}. {r['nombre']:22}  AICc = {r['aicc']:.4f}{marca}\n")
        f.write(f"\n  MODELO OPTIMIZADO EXECUTADO: {mejor['nombre']}\n\n")

        f.write("  PRONÓSTICO DE LA RED:\n")
        f.write(f"  {'Periodo':>12}  {'Pronóstico':>12}  {'LI 95%':>12}  {'LS 95%':>12}\n")
        f.write("  " + "-" * 54 + "\n")
        for a, mu, li, ls in zip(future_labels, p_mu, p_ci[:, 0], p_ci[:, 1]):
            f.write(f"  {str(a):>12}  {mu:>12.2f}  {li:>12.2f}  {ls:>12.2f}\n")

        # Ecuación matemática
        f.write("\n" + LINEA + "\n")
        f.write("  ECUACIÓN MATEMÁTICA DEL MODELO (Metodología Box-Jenkins)\n")
        f.write(LINEA + "\n\n")
        
        params = mejor["res"].params
        pnames = mejor["res"].param_names
        
        c = params[pnames.index("intercept")] if "intercept" in pnames else (params[pnames.index("const")] if "const" in pnames else 0.0)
        phi1 = params[pnames.index("ar.L1")] if "ar.L1" in pnames else 0.0
        Phi1 = params[pnames.index("ar.S.L4")] if "ar.S.L4" in pnames else 0.0
        
        def fmt_coef(val, name, first=False):
            if val == 0: return ""
            sign = "" if first else (" + " if val > 0 else " - ")
            abs_val = abs(val) if not first else val
            return f"{sign}{abs_val:.4f}{name}"

        eq_rhs = f"{c:.4f}"
        eq_rhs += fmt_coef(phi1, "(ΔΔ4 TD_t-1)")
        eq_rhs += fmt_coef(Phi1, "(ΔΔ4 TD_t-4)")
        
        f.write(f"  ΔΔ4 TD_t = {eq_rhs}\n\n")
        f.write("  Donde:\n")
        f.write("  ΔΔ4 TD_t = (TD_t - TD_t-1) - (TD_t-4 - TD_t-5)\n")

def generate_excel_report(out_dir, mejor, pron_df):
    try:
        excel_path = os.path.join(out_dir, "reporte_econometrico_profesional.xlsx")
        res = mejor["res"]
        pnames = res.param_names
        params = res.params
        
        c_val = params[pnames.index("intercept")] if "intercept" in pnames else (params[pnames.index("const")] if "const" in pnames else 0.0)
        phi1_val = params[pnames.index("ar.L1")] if "ar.L1" in pnames else 0.0
        Phi1_val = params[pnames.index("ar.S.L4")] if "ar.S.L4" in pnames else 0.0
        sigma2_val = params[pnames.index("sigma2")]
        
        resumen_data = [
            ["Parámetro / Métrica", "Valor"],
            ["Constante (Drift)", round(c_val, 6)],
            ["Rezago Regular (Y_t-1)", round(phi1_val, 6)],
            ["Rezago Estacional (Y_t-4)", round(Phi1_val, 6)],
            ["Varianza del error (sigma2)", round(sigma2_val, 6)],
            ["", ""],
            ["AICc", round(mejor["aicc"], 4)],
            ["BIC", round(mejor["bic"], 4)],
            ["RMSE", round(mejor["rmse"], 4)],
            ["p-valor Ljung-Box (Ruido Blanco)", round(mejor["p_lb"], 4)]
        ]
        df_resumen = pd.DataFrame(resumen_data[1:], columns=resumen_data[0])
        
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            df_resumen.to_excel(writer, sheet_name='Resumen Estimación', index=False)
            pron_df.to_excel(writer, sheet_name='Tabla de Pronósticos', index=False)
            
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            azul_ufps = PatternFill(start_color='1A3A6B', end_color='1A3A6B', fill_type='solid')
            gris_claro = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
            blanco_font = Font(color='FFFFFF', bold=True)
            border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

            for sheetname in writer.sheets:
                ws = writer.sheets[sheetname]
                for cell in ws[1]:
                    cell.fill = azul_ufps
                    cell.font = blanco_font
                    cell.alignment = Alignment(horizontal='center')
                    cell.border = border
                for row_idx, row in enumerate(ws.iter_rows(min_row=2), 2):
                    for cell in row:
                        cell.border = border
                        if row_idx % 2 == 0:
                            cell.fill = gris_claro
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except: pass
                    ws.column_dimensions[column_letter].width = max_length + 5
    except Exception as e:
        print(f"Nota: No se pudo generar el Excel estilizado: {e}")

def generate_plots(out_dir, variable_name, file_path, y, steps, mejor, future_labels, p_mu, p_ci, data_time_label, acf_pacf):
    fitted = mejor["res"].fittedvalues
    resid = mejor["res"].resid
    d_total = mejor["orden"][1] + (mejor["seasonal_order"][1] * 4)
    lim95 = acf_pacf["lim95"]
    acf_vals = acf_pacf["acf_vals"]
    pacf_vals = acf_pacf["pacf_vals"]

    fig = plt.figure(figsize=(16, 14), facecolor="white")
    fig.suptitle(f"Análisis SARIMA Estacional — {variable_name}\nEstructura: {mejor['nombre']} | Fuente: {os.path.basename(file_path)}", fontsize=14, fontweight="bold", color=AZUL, y=0.98)
    gs = gridspec.GridSpec(3, 2, figure=fig, hspace=0.45, wspace=0.32, top=0.92, bottom=0.07)

    ax1 = fig.add_subplot(gs[0, :])
    ax1.plot(range(len(y)), y, 'o-', color=AZUL, lw=2, ms=6, label='Histórico Original', zorder=3)
    ax1.plot(range(d_total, len(fitted)), fitted[d_total:], 's--', color=NARANJ, lw=1.8, markersize=4, label='Ajuste del Modelo', zorder=2)
    ax1.plot([len(y)-1] + list(range(len(y), len(y)+steps)), [y[-1]] + list(np.round(p_mu, 2)), 's--', color=ROJO, lw=2, ms=7, label=f'Proyección {future_labels[0]}–{future_labels[-1]}', zorder=4)
    ax1.fill_between(range(len(y), len(y)+steps), p_ci[:, 0], p_ci[:, 1], color=ROJO, alpha=0.12, label='IC 95%')
    ax1.set_title(f'Serie de tiempo original, valores ajustados y proyección ({variable_name})', fontsize=11)
    ax1.set_ylabel('Valor')
    ax1.set_xlabel(str(data_time_label))
    ax1.legend(fontsize=9, loc='upper left')
    ax1.grid(alpha=0.3)

    ax2 = fig.add_subplot(gs[1, 0])
    ax2.bar(range(d_total, len(resid)), resid[d_total:], color=[ROJO if r < 0 else AZUL for r in resid[d_total:]], alpha=0.75, edgecolor='white', linewidth=0.5)
    ax2.axhline(0, color='black', lw=1)
    ax2.axhline(+2 * np.std(resid[d_total:]), color=GRIS, ls='--', lw=1, alpha=0.7)
    ax2.axhline(-2 * np.std(resid[d_total:]), color=GRIS, ls='--', lw=1, alpha=0.7)
    ax2.set_title('Residuales del Modelo (Ruido Blanco)', fontsize=10)
    ax2.set_ylabel('Residual')
    ax2.set_xlabel('Índice')
    ax2.grid(alpha=0.2)

    ax3 = fig.add_subplot(gs[1, 1])
    resid_clean = resid[d_total:]
    acf_resid = acf(resid_clean, nlags=min(8, len(resid_clean) - 1), fft=False)
    lags_r = np.arange(1, len(acf_resid))
    bars = ax3.bar(lags_r, acf_resid[1:], color=AZUL, alpha=0.7, width=0.6)
    ax3.axhline(0, color='black', lw=1)
    ax3.axhline(+lim95, color=ROJO, ls='--', lw=1.2, label='IC 95%')
    ax3.axhline(-lim95, color=ROJO, ls='--', lw=1.2)
    for bar, val in zip(bars, acf_resid[1:]):
        if abs(val) > lim95:
            bar.set_color(ROJO); bar.set_alpha(0.9)
    ax3.set_title('FAC (ACF) de los Residuales', fontsize=10)
    ax3.set_xlabel('Lag'); ax3.set_ylabel('Autocorrelación'); ax3.set_xticks(lags_r); ax3.legend(fontsize=8); ax3.grid(alpha=0.2)

    ax4l = fig.add_subplot(gs[2, 0])
    ax4r = fig.add_subplot(gs[2, 1])
    lags_w = np.arange(1, len(acf_vals) + 1)
    def barras_acf(ax, valores, titulo, color_base):
        bars = ax.bar(lags_w, valores, color=color_base, alpha=0.7, width=0.6)
        ax.axhline(0, color='black', lw=1)
        ax.axhline(+lim95, color=ROJO, ls='--', lw=1.2)
        ax.axhline(-lim95, color=ROJO, ls='--', lw=1.2)
        for b, v in zip(bars, valores):
            if abs(v) > lim95: b.set_color(ROJO); b.set_alpha(0.9)
        ax.set_xticks(lags_w); ax.set_xlabel('Lag'); ax.set_ylabel('Autocorrelación'); ax.set_title(titulo, fontsize=10); ax.grid(alpha=0.2)
    barras_acf(ax4l, acf_vals, 'FAC de Wt = Yt − Yt−1 (Filtro Tendencial)', AZUL)
    barras_acf(ax4r, pacf_vals, 'FACP de Wt = Yt − Yt−1', VERDE)

    plt.savefig(os.path.join(out_dir, "grafico_diagnostico.png"), dpi=150, bbox_inches='tight', facecolor='white')
    plt.close()

    fig2, ax = plt.subplots(figsize=(12, 6), facecolor='white')
    ax.plot(range(len(y)), y, 'o-', lw=2, ms=7, label=f'Histórico ({os.path.basename(file_path)})', zorder=3)
    ax.plot([len(y)-1] + list(range(len(y), len(y)+steps)), [y[-1]] + list(np.round(p_mu, 2)), 's--', lw=2.5, ms=8, label=f'Proyección {mejor["nombre"]}', zorder=4)
    ax.fill_between(range(len(y), len(y)+steps), p_ci[:, 0], p_ci[:, 1], alpha=0.12, label='Intervalo de confianza 95%')
    for a, mu in zip(range(len(y), len(y)+steps), p_mu):
        ax.annotate(f"{mu:.1f}", (a, mu), textcoords='offset points', xytext=(0, 10), ha='center', fontsize=9, fontweight='bold')
    ax.set_title(f'Proyección Trimestral Estacional: {variable_name}\nModelo {mejor["nombre"]}', fontsize=13, fontweight='bold')
    ax.set_xlabel('Índice temporal'); ax.set_ylabel('Valor'); ax.legend(fontsize=10, loc='upper left'); ax.grid(alpha=0.3)
    plt.tight_layout()
    plt.savefig(os.path.join(out_dir, "grafico_pronostico.png"), dpi=150, bbox_inches='tight', facecolor='white')
    plt.close()
