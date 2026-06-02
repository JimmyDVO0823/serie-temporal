"""
ARIMA/SARIMA interactivo y universal - Versión Especial Modelo UFPS
- Lee formatos horizontales y verticales (CSV, XLSX, XLS)
- Incorpora rezagos estacionales (Y_t-4) y diferencias estacionales (D=1)
- Genera reporte econométrico completo, pronóstico y gráficas de diagnóstico

Formato vertical:
    fila 1 = años
    columna 1 = variables

Formato horizontal:
    columnas = variables / tiempo
    filas = periodos
"""

import os
import re
import glob
import warnings
import argparse
from pathlib import Path

import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
from statsmodels.tsa.statespace.sarimax import SARIMAX
from statsmodels.tsa.stattools import adfuller, acf, pacf
from statsmodels.stats.diagnostic import acorr_ljungbox
from sklearn.metrics import mean_squared_error, mean_absolute_error

warnings.filterwarnings("ignore")

# ---------------------------------------------------------------------
# Utilidades de entrada
# ---------------------------------------------------------------------

TIME_COLS = {
    "año", "anio", "year", "time", "tiempo", "t", "trimestre", "quarter",
    "fecha", "date", "periodo", "period", "mes", "month"
}

def ask_choice(prompt, options, default=None):
    if not options:
        raise ValueError("No hay opciones para elegir.")
    while True:
        print("\n" + prompt)
        for i, opt in enumerate(options, 1):
            print(f"  {i}. {opt}")
        if default is not None:
            raw = input(f"Elige una opción [1-{len(options)}] (Enter = {default}): ").strip()
        else:
            raw = input(f"Elige una opción [1-{len(options)}]: ").strip()

        if raw == "" and default is not None:
            return options[default - 1]

        try:
            idx = int(raw)
            if 1 <= idx <= len(options):
                return options[idx - 1]
        except ValueError:
            pass
        print("Entrada inválida. Intenta de nuevo.")

def ask_int(prompt, default=3, min_value=1):
    while True:
        raw = input(f"{prompt} (Enter = {default}): ").strip()
        if raw == "":
            return default
        try:
            val = int(raw)
            if val >= min_value:
                return val
        except ValueError:
            pass
        print(f"Ingresa un número entero >= {min_value}.")

def _parse_number_cell(v, default_mode="auto"):
    if pd.isna(v):
        return np.nan
    s = str(v).strip()
    if s in {"", "nan", "NaN", "-", "—"}:
        return np.nan

    if default_mode == "dot_thousands":
        s = s.replace(".", "").replace(",", ".")
    elif default_mode == "comma_decimal":
        s = s.replace(",", ".")
    elif default_mode == "dot_decimal":
        if "," in s and "." in s:
            if s.rfind(",") > s.rfind("."):
                s = s.replace(".", "").replace(",", ".")
            else:
                s = s.replace(",", "")
        elif "," in s:
            s = s.replace(",", ".")
    else:
        if re.match(r"^\d{1,3}(\.\d{3})+(,\d+)?$", s):
            s = s.replace(".", "").replace(",", ".")
        elif re.match(r"^\d{1,3}(,\d{3})+(\.\d+)?$", s):
            s = s.replace(",", "")
        elif "," in s and "." in s:
            if s.rfind(",") > s.rfind("."):
                s = s.replace(".", "").replace(",", ".")
            else:
                s = s.replace(",", "")
        elif "," in s:
            s = s.replace(",", ".")
    return pd.to_numeric(s, errors="coerce")

def normalize_numeric_series(s):
    ser = pd.Series(s, dtype="object")
    txt = ser.astype(str).str.strip()

    non_empty = txt[~txt.isin(["", "nan", "NaN", "-", "—"])]

    if len(non_empty) == 0:
        return ser.map(_parse_number_cell)

    has_comma = non_empty.str.contains(",", regex=False).any()
    has_dot = non_empty.str.contains(".", regex=False).any()

    dot_thousands = non_empty.str.match(r"^\d{1,3}(\.\d{3})+$", na=False).sum()

    if has_comma and not has_dot:
        mode = "comma_decimal"
    elif has_dot and not has_comma:
        mode = "dot_thousands" if dot_thousands >= max(2, len(non_empty) // 2) else "dot_decimal"
    else:
        mode = "auto"

    return txt.map(lambda v: _parse_number_cell(v, default_mode=mode))

def clean_name(name):
    name = str(name)
    name = re.sub(r"[^\w\s\-]+", "", name, flags=re.UNICODE)
    name = re.sub(r"\s+", "_", name.strip())
    return name[:80] if name else "serie"

def read_file_flexible(path):
    ext = Path(path).suffix.lower()
    if ext in [".xlsx", ".xls"]:
        try:
            df = pd.read_excel(path, header=None)
            return df, "excel"
        except Exception as e:
            raise RuntimeError(f"Error leyendo el archivo Excel: {e}")
    else:
        encodings = ["utf-8", "utf-8-sig", "latin1", "cp1252"]
        last_error = None
        for enc in encodings:
            try:
                try:
                    df = pd.read_csv(path, header=None, sep=None, engine="python", encoding=enc)
                except Exception:
                    df = pd.read_csv(path, header=0, sep=None, engine="python", encoding=enc)
                return df, enc
            except Exception as e:
                last_error = e
        raise last_error

def is_year_like(v):
    try:
        n = int(float(str(v).replace(",", ".").strip()))
        return 1900 <= n <= 2100
    except Exception:
        return False

def detect_format(path):
    df0, _ = read_file_flexible(path)
    first_row = df0.iloc[0].tolist() if len(df0) else []
    year_like = sum(is_year_like(x) for x in first_row)
    first_col_text = sum(not pd.isna(x) and not is_year_like(x) for x in df0.iloc[:, 0].tolist()) if df0.shape[1] else 0
    if year_like >= max(4, len(first_row) // 2):
        return "vertical"
    
    ext = Path(path).suffix.lower()
    if ext in [".xlsx", ".xls"]:
        try:
            dfh = pd.read_excel(path)
        except Exception:
            dfh = None
    else:
        try:
            dfh = pd.read_csv(path, sep=None, engine="python", encoding="utf-8")
        except Exception:
            try:
                dfh = pd.read_csv(path, sep=None, engine="python", encoding="latin1")
            except Exception:
                dfh = None
                
    if dfh is not None and any(str(c).strip().lower() in TIME_COLS for c in dfh.columns):
        return "horizontal"
    if first_col_text > 3 and year_like < 3:
        return "horizontal"
    return "vertical"

def list_supported_files():
    extensions = ("*.csv", "*.CSV", "*.xlsx", "*.XLSX", "*.xls", "*.XLS")
    files = []
    seen = set()
    for ext in extensions:
        for p in sorted(glob.glob(ext)):
            if p not in seen:
                seen.add(p)
                files.append(p)
    return files

# ---------------------------------------------------------------------
# Carga de datos
# ---------------------------------------------------------------------

def choose_file_interactive():
    files = list_supported_files()
    if not files:
        raise FileNotFoundError("No se encontraron archivos CSV o Excel (.xlsx, .xls) en el directorio actual.")
    if len(files) == 1:
        print(f"Archivo detectado: {files[0]}")
        return files[0]
    return ask_choice("Selecciona el archivo para analizar:", files)

def load_vertical(path):
    df, enc = read_file_flexible(path)
    row_years = None
    for idx, row in df.iterrows():
        nums = normalize_numeric_series(row)
        years = nums.dropna()
        years = years[(years >= 1900) & (years <= 2100)]
        if len(years) >= 4:
            row_years = idx
            break
    if row_years is None:
        raise ValueError("No se pudo detectar la fila de años.")

    years_raw = normalize_numeric_series(df.iloc[row_years]).dropna()
    years = years_raw[(years_raw >= 1900) & (years_raw <= 2100)].astype(int).tolist()

    variables = []
    map_idx = []
    for idx, val in df.iloc[:, 0].items():
        if idx == row_years:
            continue
        if pd.isna(val):
            continue
        name = str(val).strip()
        if name and name.lower() != "nan":
            variables.append(name)
            map_idx.append(idx)

    if not variables:
        raise ValueError("No se detectaron variables en la primera columna.")

    chosen = ask_choice("Variables detectadas en el archivo:", variables)
    idx_var = map_idx[variables.index(chosen)]

    raw_vals = df.iloc[idx_var, 1:1 + len(years)]
    y = normalize_numeric_series(raw_vals).to_numpy(dtype=float)
    years_arr = np.array(years, dtype=int)

    mask = ~np.isnan(y)
    y = y[mask]
    years_arr = years_arr[mask].tolist()

    return {
        "format": "vertical",
        "file": path,
        "encoding": enc,
        "variable_name": chosen,
        "series": y,
        "time": years_arr,
        "time_label": "Año",
        "time_mode": "year",
    }

def quarter_to_num(q):
    q = str(q).strip().upper()
    mapping = {
        "I": 1, "1": 1, "Q1": 1,
        "II": 2, "2": 2, "Q2": 2,
        "III": 3, "3": 3, "Q3": 3,
        "IV": 4, "4": 4, "Q4": 4,
    }
    return mapping.get(q)

def build_quarter_labels(years, quarters):
    labels = []
    for y, q in zip(years, quarters):
        qn = quarter_to_num(q)
        if qn is None:
            labels.append(f"{y} {q}")
        else:
            labels.append(f"{int(y)} Q{qn}")
    return labels

def next_quarters(last_year, last_quarter_num, steps):
    out = []
    y, q = int(last_year), int(last_quarter_num)
    for _ in range(steps):
        q += 1
        if q > 4:
            q = 1
            y += 1
        out.append((y, q))
    return out

def load_horizontal(path):
    ext = Path(path).suffix.lower()
    if ext in [".xlsx", ".xls"]:
        df = pd.read_excel(path)
    else:
        df = pd.read_csv(path, sep=None, engine="python", encoding="utf-8")
        if df.shape[1] == 1:
            df = pd.read_csv(path, sep=None, engine="python", encoding="latin1")

    cols = list(df.columns)

    candidate_series = []
    for c in cols:
        s = normalize_numeric_series(df[c])
        non_na = s.notna().sum()
        if non_na >= max(4, len(df) // 3):
            if str(c).strip().lower() not in TIME_COLS:
                candidate_series.append(c)

    if not candidate_series:
        candidate_series = [c for c in cols if str(c).strip().lower() not in TIME_COLS]

    chosen_series = ask_choice("Columnas/variables detectadas para analizar:", candidate_series)

    time_options = []
    has_year = any(str(c).strip().lower() in {"año", "anio", "year"} for c in cols)
    has_quarter = any(str(c).strip().lower() in {"trimestre", "quarter"} for c in cols)
    has_t = any(str(c).strip().lower() == "t" for c in cols)

    if has_year and has_quarter:
        time_options.append("Año + Trimestre")
    if has_t:
        time_options.append("t")
    if has_year:
        time_options.append("Año")
        
    extra_time_candidates = [c for c in cols if str(c).strip().lower() in TIME_COLS and c != chosen_series]
    for c in extra_time_candidates:
        if c not in time_options:
            time_options.append(c)

    if not time_options:
        time_options = ["Índice de fila"]

    chosen_time = ask_choice("Opciones de tiempo disponibles:", time_options)

    y = normalize_numeric_series(df[chosen_series]).to_numpy(dtype=float)

    year_cols = [c for c in cols if str(c).strip().lower() in {"año", "anio", "year"}]
    quarter_cols = [c for c in cols if str(c).strip().lower() in {"trimestre", "quarter"}]

    # Si el archivo trae año y trimestre, usamos una etiqueta anual-trimestral coherente.
    if chosen_time == "Año + Trimestre" or (chosen_time.lower() in {"trimestre", "quarter"} and year_cols and quarter_cols):
        years = pd.to_numeric(df[year_cols[0]], errors="coerce")
        quarters = df[quarter_cols[0]].astype(str)
        time_labels = build_quarter_labels(years, quarters)
        time_mode = "quarter"
    elif chosen_time == "t":
        tcol = [c for c in cols if str(c).strip().lower() == "t"][0]
        tvals = pd.to_numeric(df[tcol], errors="coerce")
        time_labels = tvals.tolist()
        time_mode = "numeric"
    elif chosen_time.lower() in {"año", "anio", "year"} and year_cols:
        tvals = pd.to_numeric(df[year_cols[0]], errors="coerce")
        time_labels = tvals.tolist()
        time_mode = "year"
    elif chosen_time.lower() in {"trimestre", "quarter"} and quarter_cols:
        # Solo trimestre, sin año: se usa como etiqueta simple.
        time_labels = df[quarter_cols[0]].astype(str).tolist()
        time_mode = "index"
    else:
        time_labels = list(range(1, len(df) + 1))
        time_mode = "index"

    mask = ~np.isnan(y)
    y = y[mask]
    time_labels = [time_labels[i] for i in range(len(time_labels)) if mask[i]]

    if time_mode == "quarter":
        years = years[mask].tolist()
        quarters = quarters[mask].tolist()
        time_labels = build_quarter_labels(years, quarters)

    return {
        "format": "horizontal",
        "file": path,
        "variable_name": chosen_series,
        "series": y,
        "time": time_labels,
        "time_label": chosen_time,
        "time_mode": time_mode,
        "raw_df": df,
    }

def load_data_interactive(args):
    file_path = args.file
    if not file_path:
        file_path = choose_file_interactive()

    if not Path(file_path).exists():
        raise FileNotFoundError(f"No existe el archivo: {file_path}")

    fmt = detect_format(file_path)
    print(f"\nFormato detectado: {fmt.upper()}")

    if fmt == "vertical":
        data = load_vertical(file_path)
    else:
        data = load_horizontal(file_path)

    return data

# ---------------------------------------------------------------------
# Pronóstico / utilidades SARIMA (Modelo Estacional UFPS)
# ---------------------------------------------------------------------

def make_future_labels(time_values, time_mode, steps):
    if time_mode == "year":
        last = int(time_values[-1])
        return list(range(last + 1, last + 1 + steps))
    if time_mode == "numeric":
        last = float(time_values[-1])
        if last.is_integer():
            last = int(last)
            return list(range(last + 1, last + 1 + steps))
        return [last + i + 1 for i in range(steps)]
    if time_mode == "quarter":
        m = re.match(r"(\d{4})\s*Q([1-4])", str(time_values[-1]))
        if m:
            year, q = int(m.group(1)), int(m.group(2))
            nxt = next_quarters(year, q, steps)
            return [f"{y} Q{q}" for y, q in nxt]
    start = len(time_values)
    return list(range(start + 1, start + 1 + steps))

def prepare_numeric_axis(time_values, time_mode):
    if time_mode in {"year", "numeric"}:
        return np.array(pd.to_numeric(pd.Series(time_values), errors="coerce"), dtype=float)
    return np.arange(len(time_values), dtype=float)

def run_arima_analysis(data, steps):
    y = np.asarray(data["series"], dtype=float)
    time_values = list(data["time"])
    time_mode = data["time_mode"]
    variable_name = data["variable_name"]
    file_path = data["file"]

    # SARIMA Estacional requiere un set de datos mínimo para calcular diferencias de rezago 4
    if len(y) < 12:
        raise ValueError("La serie tiene muy pocos datos para aplicar un análisis SARIMA con estacionalidad (Mínimo requerido: 12 datos).")

    x = prepare_numeric_axis(time_values, time_mode)
    
    # Diferencia Regular (d=1)
    wt = np.diff(y)
    
    n = len(y)

    out_dir = f"resultados_{clean_name(variable_name)}"
    os.makedirs(out_dir, exist_ok=True)

    adf_y = adfuller(y, maxlag=min(3, len(y)-2), autolag=None)
    adf_wt = adfuller(wt, maxlag=min(3, len(wt)-2), autolag=None)

    estac_y = adf_y[1] < 0.05
    estac_wt = adf_wt[1] < 0.05

    n_lags = min(8, max(1, len(wt)-1))
    acf_vals = acf(wt, nlags=n_lags, fft=False)[1:]
    pacf_vals = pacf(wt, nlags=n_lags)[1:]
    lim95 = 1.96 / np.sqrt(len(wt))

    # --- CONFIGURACIÓN ESTRUCTURA REQUERIDA POR EL PROFESOR (SARIMA) ---
    # Formato: (Nombre, Orden Regular (p,d,q), Orden Estacional (P,D,Q,s))
    # s = 4 representa la estructura trimestral de rezago Y_{t-4}
    models = [
        ("SARIMA(0,1,0)x(0,1,0)4", (0, 1, 0), (0, 1, 0, 4)),
        ("SARIMA(1,1,0)x(0,1,0)4", (1, 1, 0), (0, 1, 0, 4)),
        ("SARIMA(1,1,0)x(1,1,0)4", (1, 1, 0), (1, 1, 0, 4)),  # <--- MODELO EXPLICITADO POR TU PROFESOR (Y_t-1, d=1, Y_t-4, D=1)
        ("SARIMA(1,1,1)x(1,1,1)4", (1, 1, 1), (1, 1, 1, 4)),
    ]

    results = []
    scale = max(float(np.nanmax(np.abs(y))), 1.0)

    for name, order, seasonal_order in models:
        try:
            # SARIMAX sí soporta la estructura estacional (p,d,q) x (P,D,Q,s)
            res = SARIMAX(
                y,
                order=order,
                seasonal_order=seasonal_order,
                trend="c",
                enforce_stationarity=False,
                enforce_invertibility=False,
            ).fit(disp=False)
            k = len(res.params)
            denom = max(n - k - 1, 1)
            aicc = res.aic + (2 * k * (k + 1)) / denom
            
            # El índice de inicio real depende de la cantidad de pérdidas por rezagos completos
            start_idx = order[1] + (seasonal_order[1] * 4)
            
            rmse = np.sqrt(mean_squared_error(y[start_idx:], res.fittedvalues[start_idx:]))
            mae = mean_absolute_error(y[start_idx:], res.fittedvalues[start_idx:])
            lb_lag = min(10, max(1, len(res.resid[start_idx:]) - 1))
            lb = acorr_ljungbox(res.resid[start_idx:], lags=[lb_lag], return_df=True)
            p_lb = float(lb["lb_pvalue"].values[0])

            ma_ok = True
            if "ma.L1" in res.param_names:
                theta = res.params[res.param_names.index("ma.L1")]
                if abs(theta) >= 0.99:
                    ma_ok = False
            if "ma.S.L4" in res.param_names:
                theta_s = res.params[res.param_names.index("ma.S.L4")]
                if abs(theta_s) >= 0.99:
                    ma_ok = False

            fc_test = np.asarray(res.get_forecast(steps=min(steps, 3)).predicted_mean)
            forecast_ok = np.all(np.isfinite(fc_test))
            if forecast_ok and np.any(np.abs(fc_test) > scale * 50):
                forecast_ok = False

            results.append({
                "nombre": name,
                "orden": order,
                "seasonal_order": seasonal_order,
                "res": res,
                "aic": res.aic,
                "aicc": aicc,
                "bic": res.bic,
                "rmse": rmse,
                "mae": mae,
                "p_lb": p_lb,
                "ma_ok": ma_ok,
                "forecast_ok": forecast_ok,
            })
        except Exception as e:
            results.append({"nombre": name, "orden": order, "res": None, "error": str(e)})

    validos = [
        r for r in results
        if r.get("res") is not None and r.get("ma_ok", True) and r.get("forecast_ok", True)
    ]

    if not validos:
        validos = [r for r in results if r.get("res") is not None and r.get("ma_ok", True)]

    if not validos:
        raise RuntimeError("No se pudo estimar ningún modelo SARIMA válido con estos datos.")

    mejor = min(validos, key=lambda x: x["aicc"])

    fc = mejor["res"].get_forecast(steps=steps)
    p_mu = np.asarray(fc.predicted_mean)
    p_ci = np.asarray(fc.conf_int(alpha=0.05))
    future_labels = make_future_labels(time_values, time_mode, steps)

    LINEA = "=" * 60
    # Mapeo econométrico adaptado a la teoría de Box-Jenkins
    mapping = {
        "intercept": "Intercepto / Drift (c)",
        "const": "Intercepto / Drift (c)",
        "ar.L1": "phi1  (AR regular lag 1 - Y_t-1)",
        "ar.L2": "phi2  (AR regular lag 2)",
        "ma.L1": "theta1 (MA regular lag 1)",
        "ar.S.L4": "Phi1 (SAR estacional lag 4 - Y_t-4)",
        "ma.S.L4": "Theta1 (SMA estacional lag 4)",
        "sigma2": "sigma2 (Varianza del error)",
    }

    report_path = os.path.join(out_dir, "reporte_modelos.txt")
    with open(report_path, "w", encoding="utf-8") as f:
        f.write(LINEA + "\n")
        f.write(f"  REPORTE SARIMA ESTACIONAL — {variable_name.upper()}\n")
        f.write(f"  Archivo: {file_path}\n")
        f.write("  Estructura Completa: [Y_t, Y_t-1, d=1, Y_t-4, D=1]\n")
        f.write(LINEA + "\n\n")

        f.write("── PRUEBA DE ESTACIONARIEDAD (ADF) ──────────────────────\n")
        f.write(f"Serie Yt (Original): estadístico ADF = {adf_y[0]:.4f} | p-valor = {adf_y[1]:.4f}\n")
        f.write(f"  → {'Estacionaria' if estac_y else 'NO estacionaria → Requiere diferencia regular (d=1)'}\n\n")
        f.write(f"Serie Wt (Diff Regular): estadístico ADF = {adf_wt[0]:.4f} | p-valor = {adf_wt[1]:.4f}\n")
        f.write(f"  → {'Estacionaria ✓ (d=1 confirmado)' if estac_wt else 'Sugerencia: Revisar combinación con D=1'}\n\n")

        f.write("── ACF / PACF DE Wt (límite ±{:.3f}) ──────────────────\n".format(lim95))
        f.write(f"  {'Lag':<5} {'ACF':>8} {'PACF':>8}  Significativo\n")
        for i in range(len(acf_vals)):
            sig = "SIG" if abs(acf_vals[i]) > lim95 or abs(pacf_vals[i]) > lim95 else "ns"
            f.write(f"  {i+1:<5} {acf_vals[i]:>+8.4f} {pacf_vals[i]:>+8.4f}  {sig}\n")
        f.write("\n")

        f.write(LINEA + "\n")
        f.write("  ESTIMACIÓN DE MODELOS CANDIDATOS\n")
        f.write(LINEA + "\n\n")

        for r in results:
            if r.get("res") is None:
                f.write(f">> {r['nombre']}  —  ERROR: {r.get('error','?')}\n\n")
                continue
            res = r["res"]
            alerta = "  ⚠ Unidades cerca del límite (descartado por invertibilidad)" if not r["ma_ok"] else ""
            f.write(f">> {r['nombre']}{alerta}\n")
            f.write("-" * 50 + "\n")
            f.write(f"  AIC  = {r['aic']:.4f}\n")
            f.write(f"  AICc = {r['aicc']:.4f}\n")
            f.write(f"  BIC  = {r['bic']:.4f}\n")
            f.write(f"  RMSE = {r['rmse']:.4f}\n")
            f.write(f"  MAE  = {r['mae']:.4f}\n")
            f.write(f"  Ljung-Box p-valor = {r['p_lb']:.4f}\n")
            f.write("\n  Parámetros Estimados:\n")
            f.write(f"  {'Parámetro':<22} {'Coeficiente':>12} {'p-valor':>10}  Signif.\n")
            f.write("  " + "-" * 55 + "\n")
            for pname in res.param_names:
                if pname == "sigma2":
                    continue
                coef = res.params[res.param_names.index(pname)]
                pval = res.pvalues[res.param_names.index(pname)]
                disp = mapping.get(pname, pname)
                sig = "✓ Signif." if pval < 0.05 else "✗ No signif."
                f.write(f"  {disp:<22} {coef:>12.6f} {pval:>10.4f}  {sig}\n")
            sigma2 = res.params[res.param_names.index("sigma2")]
            f.write(f"  {'sigma2 (varianza error)':<22} {sigma2:>12.6f}\n\n")

        f.write(LINEA + "\n")
        f.write("  SELECCIÓN DEL MEJOR MODELO ECONÓMETRICO\n")
        f.write(LINEA + "\n\n")
        f.write("  Criterio de decisión: Mínimo AICc (corregido para muestras pequeñas).\n\n")
        f.write("  Ranking General:\n")
        for i, r in enumerate(sorted(validos, key=lambda x: x["aicc"]), 1):
            marca = " ← SELECCIONADO" if r["nombre"] == mejor["nombre"] else ""
            f.write(f"    {i}. {r['nombre']:22}  AICc = {r['aicc']:.4f}{marca}\n")
        f.write(f"\n  MODELO OPTIMIZADO EXECUTADO: {mejor['nombre']}\n\n")

        f.write("  PRONÓSTICO DE LA RED:\n")
        f.write(f"  {'Periodo':>12}  {'Pronóstico':>12}  {'LI 95%':>12}  {'LS 95%':>12}\n")
        f.write("  " + "-" * 54 + "\n")
        for a, mu, li, ls in zip(future_labels, p_mu, p_ci[:, 0], p_ci[:, 1]):
            f.write(f"  {str(a):>12}  {mu:>12.2f}  {li:>12.2f}  {ls:>12.2f}\n")

        # --- ADICIÓN: ECUACIÓN MATEMÁTICA FORMAL (BOX-JENKINS) ---
        f.write("\n" + LINEA + "\n")
        f.write("  ECUACIÓN MATEMÁTICA DEL MODELO (Metodología Box-Jenkins)\n")
        f.write(LINEA + "\n\n")
        
        # Obtener coeficientes para la ecuación del modelo SARIMA(1,1,0)x(1,1,0)4
        params = mejor["res"].params
        pnames = mejor["res"].param_names
        
        c = params[pnames.index("intercept")] if "intercept" in pnames else (params[pnames.index("const")] if "const" in pnames else 0.0)
        phi1 = params[pnames.index("ar.L1")] if "ar.L1" in pnames else 0.0
        Phi1 = params[pnames.index("ar.S.L4")] if "ar.S.L4" in pnames else 0.0
        
        def fmt_coef(val, name, first=False):
            if val == 0: return ""
            sign = "" if first else (" + " if val > 0 else " - ")
            abs_val = abs(val) if not first else val
            return f"{sign}{abs_val:.4f}{name}"

        eq_rhs = f"{c:.4f}"
        eq_rhs += fmt_coef(phi1, "(ΔΔ4 TD_t-1)")
        eq_rhs += fmt_coef(Phi1, "(ΔΔ4 TD_t-4)")
        
        f.write(f"  ΔΔ4 TD_t = {eq_rhs}\n\n")
        f.write("  Donde:\n")
        f.write("  ΔΔ4 TD_t = (TD_t - TD_t-1) - (TD_t-4 - TD_t-5)\n")

    pron_df = pd.DataFrame({
        "Periodo": future_labels,
        "Pronóstico": np.round(p_mu, 2),
        "LI 95%": np.round(p_ci[:, 0], 2),
        "LS 95%": np.round(p_ci[:, 1], 2),
    })
    pron_path = os.path.join(out_dir, f"pronostico_{clean_name(variable_name)}.csv")
    pron_df.to_csv(pron_path, index=False, encoding="utf-8")

    # --- ADICIÓN: EXCEL PROFESIONAL CON ESTILOS ---
    try:
        excel_path = os.path.join(out_dir, "reporte_econometrico_profesional.xlsx")
        
        # Preparar datos para Tabla 1: Resumen
        res = mejor["res"]
        pnames = res.param_names
        params = res.params
        
        c_val = params[pnames.index("intercept")] if "intercept" in pnames else (params[pnames.index("const")] if "const" in pnames else 0.0)
        phi1_val = params[pnames.index("ar.L1")] if "ar.L1" in pnames else 0.0
        Phi1_val = params[pnames.index("ar.S.L4")] if "ar.S.L4" in pnames else 0.0
        sigma2_val = params[pnames.index("sigma2")]
        
        resumen_data = [
            ["Parámetro / Métrica", "Valor"],
            ["Constante (Drift)", round(c_val, 6)],
            ["Rezago Regular (Y_t-1)", round(phi1_val, 6)],
            ["Rezago Estacional (Y_t-4)", round(Phi1_val, 6)],
            ["Varianza del error (sigma2)", round(sigma2_val, 6)],
            ["", ""],
            ["AICc", round(mejor["aicc"], 4)],
            ["BIC", round(mejor["bic"], 4)],
            ["RMSE", round(mejor["rmse"], 4)],
            ["p-valor Ljung-Box (Ruido Blanco)", round(mejor["p_lb"], 4)]
        ]
        df_resumen = pd.DataFrame(resumen_data[1:], columns=resumen_data[0])
        
        # Tabla 2: Pronósticos (ya tenemos pron_df)
        
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            df_resumen.to_excel(writer, sheet_name='Resumen Estimación', index=False)
            pron_df.to_excel(writer, sheet_name='Tabla de Pronósticos', index=False)
            
            # Estilización con openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
            
            azul_ufps = PatternFill(start_color='1A3A6B', end_color='1A3A6B', fill_type='solid')
            gris_claro = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
            blanco_font = Font(color='FFFFFF', bold=True)
            border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                           top=Side(style='thin'), bottom=Side(style='thin'))

            for sheetname in writer.sheets:
                ws = writer.sheets[sheetname]
                # Estilo cabecera
                for cell in ws[1]:
                    cell.fill = azul_ufps
                    cell.font = blanco_font
                    cell.alignment = Alignment(horizontal='center')
                    cell.border = border
                
                # Estilo cuerpo y auto-ajuste
                for row_idx, row in enumerate(ws.iter_rows(min_row=2), 2):
                    for cell in row:
                        cell.border = border
                        if row_idx % 2 == 0:
                            cell.fill = gris_claro
                
                # Ajustar ancho de columnas
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except: pass
                    ws.column_dimensions[column_letter].width = max_length + 5

    except Exception as e:
        print(f"Nota: No se pudo generar el Excel estilizado (requiere openpyxl): {e}")

    AZUL = "#1A3A6B"
    ROJO = "#C0392B"
    VERDE = "#27AE60"
    NARANJ = "#E67E22"
    GRIS = "#7F8C8D"

    fitted = mejor["res"].fittedvalues
    resid = mejor["res"].resid
    
    # Pérdida combinada por la diferenciación regular y estacional (d + D*s)
    d_total = mejor["orden"][1] + (mejor["seasonal_order"][1] * 4)

    fig = plt.figure(figsize=(16, 14), facecolor="white")
    fig.suptitle(
        f"Análisis SARIMA Estacional — {variable_name}\n"
        f"Estructura: {mejor['nombre']} | Fuente: {os.path.basename(file_path)}",
        fontsize=14, fontweight="bold", color=AZUL, y=0.98
    )

    gs = gridspec.GridSpec(3, 2, figure=fig, hspace=0.45, wspace=0.32, top=0.92, bottom=0.07)

    ax1 = fig.add_subplot(gs[0, :])
    ax1.plot(range(len(y)), y, 'o-', color=AZUL, lw=2, ms=6, label='Histórico Original', zorder=3)
    ax1.plot(range(d_total, len(fitted)), fitted[d_total:], 's--', color=NARANJ, lw=1.8, markersize=4, label='Ajuste del Modelo', zorder=2)
    ax1.plot([len(y)-1] + list(range(len(y), len(y)+steps)), [y[-1]] + list(np.round(p_mu, 2)), 's--', color=ROJO, lw=2, ms=7, label=f'Proyección {future_labels[0]}–{future_labels[-1]}', zorder=4)
    ax1.fill_between(range(len(y), len(y)+steps), p_ci[:, 0], p_ci[:, 1], color=ROJO, alpha=0.12, label='IC 95%')
    ax1.set_title(f'Serie de tiempo original, valores ajustados y proyección ({variable_name})', fontsize=11)
    ax1.set_ylabel('Valor')
    ax1.set_xlabel(str(data["time_label"]))
    ax1.legend(fontsize=9, loc='upper left')
    ax1.grid(alpha=0.3)

    ax2 = fig.add_subplot(gs[1, 0])
    ax2.bar(range(d_total, len(resid)), resid[d_total:], color=[ROJO if r < 0 else AZUL for r in resid[d_total:]], alpha=0.75, edgecolor='white', linewidth=0.5)
    ax2.axhline(0, color='black', lw=1)
    ax2.axhline(+2 * np.std(resid[d_total:]), color=GRIS, ls='--', lw=1, alpha=0.7)
    ax2.axhline(-2 * np.std(resid[d_total:]), color=GRIS, ls='--', lw=1, alpha=0.7)
    ax2.set_title('Residuales del Modelo (Ruido Blanco)', fontsize=10)
    ax2.set_ylabel('Residual')
    ax2.set_xlabel('Índice')
    ax2.grid(alpha=0.2)

    ax3 = fig.add_subplot(gs[1, 1])
    acf_resid = acf(resid[d_total:], nlags=min(8, len(resid[d_total:]) - 1), fft=False)
    lags_r = np.arange(1, len(acf_resid))
    bars = ax3.bar(lags_r, acf_resid[1:], color=AZUL, alpha=0.7, width=0.6)
    ax3.axhline(0, color='black', lw=1)
    ax3.axhline(+lim95, color=ROJO, ls='--', lw=1.2, label='IC 95%')
    ax3.axhline(-lim95, color=ROJO, ls='--', lw=1.2)
    for bar, val in zip(bars, acf_resid[1:]):
        if abs(val) > lim95:
            bar.set_color(ROJO)
            bar.set_alpha(0.9)
    ax3.set_title('FAC (ACF) de los Residuales', fontsize=10)
    ax3.set_xlabel('Lag')
    ax3.set_ylabel('Autocorrelación')
    ax3.set_xticks(lags_r)
    ax3.legend(fontsize=8)
    ax3.grid(alpha=0.2)

    ax4l = fig.add_subplot(gs[2, 0])
    ax4r = fig.add_subplot(gs[2, 1])
    lags_w = np.arange(1, len(acf_vals) + 1)

    def barras_acf(ax, valores, titulo, color_base):
        bars = ax.bar(lags_w, valores, color=color_base, alpha=0.7, width=0.6)
        ax.axhline(0, color='black', lw=1)
        ax.axhline(+lim95, color=ROJO, ls='--', lw=1.2)
        ax.axhline(-lim95, color=ROJO, ls='--', lw=1.2)
        for b, v in zip(bars, valores):
            if abs(v) > lim95:
                b.set_color(ROJO)
                b.set_alpha(0.9)
        ax.set_xticks(lags_w)
        ax.set_xlabel('Lag')
        ax.set_ylabel('Autocorrelación')
        ax.set_title(titulo, fontsize=10)
        ax.grid(alpha=0.2)

    barras_acf(ax4l, acf_vals, 'FAC de Wt = Yt − Yt−1 (Filtro Tendencial)', AZUL)
    barras_acf(ax4r, pacf_vals, 'FACP de Wt = Yt − Yt−1', VERDE)

    diag_path = os.path.join(out_dir, "grafico_diagnostico.png")
    plt.savefig(diag_path, dpi=150, bbox_inches='tight', facecolor='white')
    plt.close()

    fig2, ax = plt.subplots(figsize=(12, 6), facecolor='white')
    ax.plot(range(len(y)), y, 'o-', lw=2, ms=7, label=f'Histórico ({os.path.basename(file_path)})', zorder=3)
    ax.plot([len(y)-1] + list(range(len(y), len(y)+steps)), [y[-1]] + list(np.round(p_mu, 2)), 's--', lw=2.5, ms=8, label=f'Proyección {mejor["nombre"]}', zorder=4)
    ax.fill_between(range(len(y), len(y)+steps), p_ci[:, 0], p_ci[:, 1], alpha=0.12, label='Intervalo de confianza 95%')
    for a, mu in zip(range(len(y), len(y)+steps), p_mu):
        ax.annotate(f"{mu:.1f}", (a, mu), textcoords='offset points', xytext=(0, 10), ha='center', fontsize=9, fontweight='bold')
    ax.set_title(f'Proyección Trimestral Estacional: {variable_name}\nModelo {mejor["nombre"]}', fontsize=13, fontweight='bold')
    ax.set_xlabel('Índice temporal')
    ax.set_ylabel('Valor')
    ax.legend(fontsize=10, loc='upper left')
    ax.grid(alpha=0.3)

    forecast_path = os.path.join(out_dir, "grafico_pronostico.png")
    plt.tight_layout()
    plt.savefig(forecast_path, dpi=150, bbox_inches='tight', facecolor='white')
    plt.close()

    print("\n" + "=" * 55)
    print(f"  MEJOR MODELO SELECCIONADO: {mejor['nombre']}")
    print(f"  AICc = {mejor['aicc']:.4f}   BIC = {mejor['bic']:.4f}")
    print(f"  RMSE = {mejor['rmse']:.4f}   MAE = {mejor['mae']:.4f}")
    print(f"  Ljung-Box p-valor = {mejor['p_lb']:.4f}")
    print("\n  PRONÓSTICO DE PROYECCIÓN:")
    for a, mu, li, ls in zip(future_labels, p_mu, p_ci[:, 0], p_ci[:, 1]):
        print(f"    {a}: {mu:.2f}  [IC 95%: {li:.2f} – {ls:.2f}]")
    print("=" * 55)
    print(f"\n  Archivos econométricos generados con éxito en: ./{out_dir}/")
    print(f"   • reporte_modelos.txt (Muestra los parámetros analíticos de Y_t-1 y Y_t-4)")
    print(f"   • {os.path.basename(pron_path)}")
    print(f"   • grafico_diagnostico.png")
    print(f"   • grafico_pronostico.png")

    return out_dir

def parse_args():
    parser = argparse.ArgumentParser(description="SARIMA universal interactivo - Metodología UFPS")
    parser.add_argument("--file", type=str, default="", help="Ruta al archivo CSV o Excel (opcional)")
    parser.add_argument("--variable", type=str, default="", help="Variable (opcional)")
    parser.add_argument("--steps", type=int, default=3, help="Cantidad de pasos a pronosticar")
    return parser.parse_args()

def main():
    args = parse_args()
    try:
        data = load_data_interactive(args)
        steps = args.steps if args.steps and args.steps > 0 else ask_int("¿Cuántos pasos trimestrales quieres proyectar?", default=4, min_value=1)
        run_arima_analysis(data, steps)
    except Exception as e:
        print(f"\n⚠ ERROR EN SISTEMA: {e}")
        raise

if __name__ == "__main__":
    main()